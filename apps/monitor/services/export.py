"""
apps/monitor/services/export.py
-------------------------------------
Service xuất báo cáo Excel cho các loại log (audit, mailbox, metric_detail).
Dùng thư viện openpyxl để tạo file .xlsx với formatting đẹp mắt + freeze header.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


class ExcelExportService:
    """Service chung xuất Excel cho mọi loại log."""

    @staticmethod
    def _setup_worksheet_style(ws, headers, title=None):
        """
        Cài đặt style chung cho worksheet:
        - Đóng băng header (freeze panes)
        - Màu nền header xanh đậm
        - Font chữ mô tả
        - Border cho tất cả ô
        """
        # Đóng băng hàng đầu tiên (header)
        ws.freeze_panes = 'A2'

        # Style header (hàng 1)
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Style dữ liệu (từ hàng 2 trở đi) -- border + alignment
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border

    @staticmethod
    def _auto_adjust_columns(ws, headers):
        """Tự động điều chỉnh độ rộng cột theo nội dung."""
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            max_length = len(str(header))

            # Duyệt qua từng ô trong cột để tìm max_length
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)  # Giới hạn max 50 để không quá rộng
            ws.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def export_audit_logs(items, filename_prefix='audit_logs'):
        """
        Xuất danh sách audit logs sang Excel.
        items: danh sách dict từ AuditService.query()['items']
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Audit Logs'

        headers = [
            'Thời gian (@timestamp)',
            'Người thực hiện (admin_email)',
            'Email xác thực (auth_email)',
            'Hành động (command)',
            'Phân loại (action_category)',
            'Đối tượng (target_email)',
            'IP nguồn (client_ip)',
            'Trạng thái (login_status)',
            'Loại lỗi (error_type)',
            'Host',
            'Message (gốc)',
        ]

        # Ghi dữ liệu
        ws.append(headers)
        for item in items:
            row = [
                item.get('timestamp') or item.get('@timestamp', '-'),
                item.get('admin_email', '-'),
                item.get('auth_email', '-'),
                item.get('command', '-'),
                item.get('action_category', '-'),
                item.get('target_email', '-'),
                item.get('client_ip', '-'),
                item.get('login_status', '-'),
                item.get('error_type', '-'),
                item.get('host', {}).get('hostname') if isinstance(item.get('host'), dict) else '-',
                item.get('message', '-'),
            ]
            ws.append(row)

        ExcelExportService._setup_worksheet_style(ws, headers)
        ExcelExportService._auto_adjust_columns(ws, headers)

        # Lưu vào BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, f'{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    @staticmethod
    def export_mailbox_logs(items, filename_prefix='mailbox_logs'):
        """
        Xuất danh sách mailbox logs sang Excel.
        items: danh sách dict từ MailboxService.query()['items']
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mailbox Logs'

        headers = [
            'Thời gian (@timestamp)',
            'Từ (from)',
            'Đến (to)',
            'Chiều thư (mail_direction)',
            'Trạng thái (status)',
            'Kích thước (size_bytes)',
            'Độ trễ (delay_seconds)',
            'Queue ID',
            'Message ID',
            'Host',
            'Amavis Status',
        ]

        # Ghi dữ liệu
        ws.append(headers)
        for item in items:
            row = [
                item.get('timestamp') or item.get('@timestamp', '-'),
                item.get('from', '-'),
                item.get('to', '-'),
                item.get('mail_direction', '-'),
                item.get('status', '-'),
                item.get('size_bytes', '-'),
                item.get('delay_seconds', '-'),
                item.get('queue_id', '-'),
                item.get('message_id', '-'),
                item.get('host', {}).get('hostname') if isinstance(item.get('host'), dict) else '-',
                item.get('amavis_status', '-'),
            ]
            ws.append(row)

        ExcelExportService._setup_worksheet_style(ws, headers)
        ExcelExportService._auto_adjust_columns(ws, headers)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, f'{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    @staticmethod
    def export_backup_logs(items, filename_prefix='backup_logs'):
        """
        Xuất danh sách backup logs sang Excel.
        items: danh sách dict từ BackupService.query()['items']
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Backup Logs'

        headers = [
            'Thời gian (@timestamp)',
            'Tài khoản (account)',
            'Loại Backup (backup_mode)',
            'Trạng thái (status)',
            'Kích thước (size_bytes)',
            'Thời gian xử lý (duration_seconds)',
            'Host',
            'Backup Path',
        ]

        # Ghi dữ liệu
        ws.append(headers)
        for item in items:
            row = [
                item.get('timestamp') or item.get('@timestamp', '-'),
                item.get('account', '-'),
                item.get('backup_mode', '-'),
                item.get('status', '-'),
                item.get('size_bytes', '-'),
                item.get('duration_seconds', '-'),
                item.get('host', {}).get('hostname') if isinstance(item.get('host'), dict) else '-',
                item.get('backup_path', '-'),
            ]
            ws.append(row)

        ExcelExportService._setup_worksheet_style(ws, headers)
        ExcelExportService._auto_adjust_columns(ws, headers)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, f'{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    @staticmethod
    def export_metric_logs(items, hostname='unknown', filename_prefix='metric_logs'):
        """
        Xuất danh sách metric logs (chi tiết host) sang Excel.
        items: danh sách dict từ MetricService.query_paginated()['items']
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Metric Logs'

        headers = [
            'Thời điểm (@timestamp)',
            'CPU (%)',
            'RAM (%)',
            'Disk (%)',
            'RX Bitrate (bps)',
            'TX Bitrate (bps)',
            'Mail Queue',
            'Zimbra Services Down',
            'Severity',
        ]

        # Ghi dữ liệu
        ws.append(headers)
        for item in items:
            zimbra_down = ', '.join(item.get('zimbra_not_running', [])) if item.get('zimbra_not_running') else 'OK'
            row = [
                item.get('timestamp') or item.get('@timestamp', '-'),
                item.get('cpu', '-'),
                item.get('ram', '-'),
                item.get('disk', '-'),
                item.get('net_rx_bps', '-'),
                item.get('net_tx_bps', '-'),
                item.get('queue', '-'),
                zimbra_down,
                item.get('severity', '-'),
            ]
            ws.append(row)

        ExcelExportService._setup_worksheet_style(ws, headers)
        ExcelExportService._auto_adjust_columns(ws, headers)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, f'{filename_prefix}_{hostname}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'